import datetime
import openpyxl
import os
import urllib.request

from dateutil import parser

from PySide6.QtCore import Qt
from PySide6.QtWidgets import QDialog, QTableWidget, QTableWidgetItem, QVBoxLayout, QHBoxLayout, QDialogButtonBox, QComboBox
from PySide6.QtGui import QPixmap, QImage, QPainter

from OpenNumismat.Collection.Import import _Import2, _InvalidDatabaseError
from OpenNumismat.Tools.DialogDecorators import storeDlgSizeDecorator
from OpenNumismat.Collection.CollectionFields import FieldTypes as Type
from OpenNumismat.Settings import Settings
from OpenNumismat import version


@storeDlgSizeDecorator
class TableDialog(QDialog):

    def __init__(self, parent, path):
        super().__init__(parent,
                         Qt.WindowCloseButtonHint | Qt.WindowSystemMenuHint)

        self.path = path

        self.setWindowTitle(self.tr("Select columns"))

        self.hlayout = QHBoxLayout()

        buttonBox = QDialogButtonBox(Qt.Horizontal)
        buttonBox.addButton(QDialogButtonBox.Ok)
        buttonBox.accepted.connect(self.accept)

        self.table = QTableWidget(self)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)

        layout = QVBoxLayout()
        layout.addLayout(self.hlayout)
        layout.addWidget(self.table)
        layout.addWidget(buttonBox)

        self.setLayout(layout)

    def comboChanged(self, _index):
        labels = []
        for i in range(1, self.hlayout.count()):
            combo = self.hlayout.itemAt(i).widget()
            labels.append(combo.currentText())
        self.table.setHorizontalHeaderLabels(labels)

        for i in range(1, self.hlayout.count()):
            combo = self.hlayout.itemAt(i).widget()
            field = combo.currentData()
            if not field:
                continue

            if field.type == Type.Date:
                for row in range(self.table.rowCount()):
                    item = self.table.item(row, i - 1)
                    val = item.text()

                    try:
                        val = parser.parse(val).date().isoformat()
                        item.setText(val)
                    except (ValueError, TypeError):
                        pass
            elif field.type in Type.ImageTypes:
                for row in range(self.table.rowCount()):
                    item = self.table.item(row, i - 1)
                    if item.data(Qt.UserRole) is not None:
                        pixmap = QPixmap.fromImage(item.data(Qt.UserRole))
                        item.setData(Qt.DecorationRole, pixmap)
                        item.setText('')
                        continue

                    fileName = item.text()
                    image = QImage()
                    if fileName.startswith('http'):
                        try:
                            # Wikipedia require any header
                            req = urllib.request.Request(fileName,
                                    headers={'User-Agent': version.AppName})
                            data = urllib.request.urlopen(req, timeout=30).read()
                            result = image.loadFromData(data)
                            if result:
                                pixmap = QPixmap.fromImage(image)
                                item.setData(Qt.DecorationRole, pixmap)
                                item.setText('')
                        except:
                            pass
                    else:
                        if not os.path.isabs(fileName):
                            fileName = os.path.join(self.path, fileName)

                        result = image.load(fileName)
                        if result:
                            pixmap = QPixmap.fromImage(image)
                            item.setData(Qt.DecorationRole, pixmap)
                            item.setText('')


class ImportExcel(_Import2):

    def __init__(self, parent=None):
        super().__init__(parent)

    @staticmethod
    def isAvailable():
        return True

    def defaultField(self, col, _combo):
        if col < 10:
            return col + 1

        return 0

    def defaultStatus(self):
        return 'owned'

    def _connect(self, src):
        try:
            book = openpyxl.load_workbook(src)
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise _InvalidDatabaseError(str(e))

        self.sheet = book.active

        self.images = {}
        for image in self.sheet._images:
            img = QImage()
            if img.loadFromData(image._data()):
                _from = image.anchor._from
                col = openpyxl.utils.get_column_letter(_from.col + 1)
                coordinate = f"{col}{_from.row + 1}"
                self.images[coordinate] = img

        self.src_path = os.path.dirname(src)
        dialog = TableDialog(self.parent(), self.src_path)

        rows = min(self.sheet.max_row, 10)

        dialog.table.setRowCount(rows)
        dialog.table.setColumnCount(self.sheet.max_column)

        for row in range(rows):
            for col in range(self.sheet.max_column):
                cell = self.sheet.cell(row + 1, col + 1)
                val = cell.value

                if val is None:
                    val = ''
                elif isinstance(val, datetime.time):
                    val = ''
                elif isinstance(val, datetime.datetime):
                    val = val.date()
                if cell.hyperlink:
                    val = cell.hyperlink.target

                item = QTableWidgetItem(str(val))

                if cell.coordinate in self.images:
                    item.setData(Qt.UserRole, self.images[cell.coordinate])

                dialog.table.setItem(row, col, item)

        dialog.hlayout.addSpacing(dialog.table.verticalHeader().width())

        self.comboBoxes = []
        for col in range(self.sheet.max_column):
            combo = QComboBox()
            combo.addItem(self.tr("<Ignore>"))
            for f in self.fields.userFields:
                if f not in self.fields.systemFields:
                    combo.addItem(f.title, f)
            combo.setCurrentIndex(self.defaultField(col, combo))
            combo.setSizeAdjustPolicy(QComboBox.AdjustToMinimumContentsLengthWithIcon)
            combo.currentIndexChanged.connect(dialog.comboChanged)
            dialog.hlayout.addWidget(combo)

            self.comboBoxes.append(combo)
        dialog.comboChanged(0)

        result = dialog.exec()
        if result == QDialog.Accepted:
            self.selected_fields = []
            self.has_title = False
            self.has_status = False
            for i in range(1, dialog.hlayout.count()):
                combo = dialog.hlayout.itemAt(i).widget()
                field = combo.currentData()
                self.selected_fields.append(field)

                if field:
                    if field.name == 'title':
                        self.has_title = True
                    elif field.name == 'status':
                        self.has_status = True

            return book

        return None

    def _getRowsCount(self, book):
        return self.sheet.max_row

    def _setRecord(self, record, row):
        for i, field in enumerate(self.selected_fields):
            if not field:
                continue

            cell = self.sheet.cell(row + 1, i + 1)
            val = cell.value
            if isinstance(val, datetime.time):
                val = ''
            elif isinstance(val, datetime.datetime):
                val = val.date()
            if isinstance(val, datetime.date):
                val = val.isoformat()
            if cell.hyperlink:
                val = cell.hyperlink.target

            if field.type == Type.Date:
                try:
                    val = parser.parse(val).date().isoformat()
                except (ValueError, TypeError):
                    val = None
            elif field.type in Type.ImageTypes:
                if cell.coordinate in self.images:
                    image = self.images[cell.coordinate]
                    val = self.__fixTransparentImage(image)
                elif val:
                    image = QImage()
                    if val.startswith('http'):
                        try:
                            # Wikipedia require any header
                            req = urllib.request.Request(val,
                                    headers={'User-Agent': version.AppName})
                            data = urllib.request.urlopen(req, timeout=30).read()
                            if image.loadFromData(data):
                                val = self.__fixTransparentImage(image)
                            else:
                                val = None
                        except:
                            val = None
                    else:
                        if os.path.isabs(val):
                            fileName = val
                        else:
                            fileName = os.path.join(self.src_path, val)

                        if image.load(fileName):
                            val = self.__fixTransparentImage(image)
                        else:
                            val = None

            record.setValue(field.name, val)

        if not self.has_title:
            record.setValue('title', self.__generateTitle(record))

        if not self.has_status:
            record.setValue('status', self.defaultStatus())

    def __fixTransparentImage(self, image):
        if image.hasAlphaChannel() and not Settings()['transparent_store']:
            # Fill transparent color if present
            color = Settings()['transparent_color']
            fixedImage = QImage(image.size(), QImage.Format_RGB32)
            fixedImage.fill(color)
            painter = QPainter(fixedImage)
            painter.drawImage(0, 0, image)
            painter.end()
        else:
            fixedImage = image

        return fixedImage

    def __generateTitle(self, record):
        title = ""
        if record.value('country'):
            title += str(record.value('country')) + ' '
        if record.value('value'):
            title += str(record.value('value')) + ' '
        if record.value('unit'):
            title += str(record.value('unit')) + ' '
        if record.value('year'):
            title += str(record.value('year')) + ' '
        if record.value('subjectshort'):
            title += '(' + str(record.value('subjectshort')) + ') '

        return title.strip()
